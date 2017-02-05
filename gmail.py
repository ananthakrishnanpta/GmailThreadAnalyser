import imaplib
import getpass
import email
import datetime
import re
import smtplib

#opening the excel sheet

from openpyxl import *
wb = load_workbook(filename='mbrs.xlsx', read_only=False)
ws = wb.active  # ws is now an IterableWorksheet

#logging in

user = raw_input("GMail :- ")
pwd = getpass.getpass()
 
mail = imaplib.IMAP4_SSL('imap.gmail.com')
mail.login(user,pwd)
mail.list()
mail.select("inbox")



for col in ws.iter_cols(min_row=1,max_col=1, max_row = 4):
    for cell in col:
        mbr = str(cell.value)
        print("%s"%(mbr))
        
        try:
            var = '(SUBJECT "Weekly status update" FROM "' + mbr +'")'
            #result, data = mail.search(None, '(SUBJECT "foss" FROM "abhinand4858@gmail.com")')
            result, data = mail.search(None, var)
            ids = data[0] # data is a list.
            id_list = ids.split() # ids is a space separated string
            latest_email_id = id_list[-1]#get the latest 
            result, data = mail.fetch(latest_email_id, "RFC822") # fetch the email body (RFC822) for the given ID
            raw_email = data[0][1] 
            email_message = email.message_from_string(raw_email)
            #print email.utils.parseaddr(email_message['From'])
            #print email_message.items()
            #write "yes" into the adjacent cell
            ws.cell(row=cell.row, column=2).value= "yes"
            
            
        except:
            
            #write "no" to adjacent cell
            print("nothing from this e-mail")
            ws.cell(row=cell.row, column=2).value= "no"
        
        
wb.save("mbrs.xlsx")
"""
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

"""
